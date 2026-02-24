"""
spreadsheet_reader.py — Unified spreadsheet reader for CSV, XLSX, and Google Sheets.
Auto-detects format by file extension or URL pattern and returns a normalized list
of LinkedIn profile URLs.
"""

import csv
import re
from pathlib import Path
from typing import Optional

# Regex to validate LinkedIn profile URLs
LINKEDIN_URL_PATTERN = re.compile(
    r"^https?://(www\.)?linkedin\.com/in/[A-Za-z0-9\-_%]+/?(\?.*)?$"
)


def _is_google_sheets_url(file_path: str) -> bool:
    """Check if the file path is a Google Sheets URL."""
    return file_path.startswith("https://docs.google.com/spreadsheets/")


def _clean_url(url: str) -> str:
    """Strip whitespace and trailing slashes from a URL."""
    url = url.strip()
    # Remove trailing slash for consistency
    if url.endswith("/"):
        url = url.rstrip("/")
    return url


def _validate_url(url: str, row: int) -> Optional[str]:
    """
    Validate that a URL is a LinkedIn profile URL.
    Returns the cleaned URL if valid, None if invalid.
    Logs a warning for invalid URLs.
    """
    cleaned = _clean_url(url)
    if not cleaned:
        return None
    if LINKEDIN_URL_PATTERN.match(cleaned):
        return cleaned
    # Be lenient: also accept URLs without protocol
    if cleaned.startswith("linkedin.com/in/") or cleaned.startswith("www.linkedin.com/in/"):
        cleaned = "https://" + cleaned
        if LINKEDIN_URL_PATTERN.match(cleaned):
            return cleaned
    print(f"  [WARNING] Row {row}: Invalid LinkedIn URL skipped: {url}")
    return None


def _find_url_column(headers: list[str]) -> Optional[int]:
    """
    Find the column index that contains LinkedIn URLs.
    Searches header names for common patterns like 'url', 'link', 'linkedin', 'profile'.
    Only matches short header-like values — skips cells that look like actual URLs.
    """
    url_keywords = ["url", "link", "linkedin", "profile", "href"]
    for i, header in enumerate(headers):
        header_lower = header.strip().lower()
        # Skip cells that look like actual URLs (not column headers)
        if header_lower.startswith("http://") or header_lower.startswith("https://"):
            continue
        for keyword in url_keywords:
            if keyword in header_lower:
                return i
    return None


def read_csv(file_path: str) -> list[dict]:
    """
    Read LinkedIn URLs from a CSV file.
    Auto-detects the URL column from headers, or uses the first column.
    """
    results = []
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"CSV file not found: {file_path}")
    if not path.suffix.lower() == ".csv":
        raise ValueError(f"Expected a .csv file, got: {path.suffix}")

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        # Sniff the dialect to handle different delimiters
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(f, dialect)
        rows = list(reader)

    if not rows:
        raise ValueError(f"CSV file is empty: {file_path}")

    # Try to detect header row
    first_row = rows[0]
    url_col = _find_url_column(first_row)

    if url_col is not None:
        # Header found — skip first row, use detected column
        data_rows = rows[1:]
        start_row = 2  # 1-indexed, accounting for header
    else:
        # No header detected — check if first row itself contains a URL
        url_col = 0
        first_cell = _clean_url(first_row[0]) if first_row else ""
        # If the first cell looks like a LinkedIn URL, treat ALL rows as data
        if LINKEDIN_URL_PATTERN.match(first_cell) or (
            first_cell.startswith("linkedin.com/in/") or first_cell.startswith("www.linkedin.com/in/")
        ):
            data_rows = rows
            start_row = 1
        else:
            # First row is an unrecognized header — skip it
            data_rows = rows[1:]
            start_row = 2

    for i, row in enumerate(data_rows):
        row_num = start_row + i
        if url_col < len(row):
            url = _validate_url(row[url_col], row_num)
            if url:
                results.append({"url": url, "row": row_num})

    return results


def read_xlsx(file_path: str) -> list[dict]:
    """
    Read LinkedIn URLs from an Excel (.xlsx) file.
    Auto-detects the URL column from headers, or uses the first column.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for .xlsx files. Install with: pip install openpyxl")

    results = []
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    if path.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raise ValueError(f"Expected an Excel file, got: {path.suffix}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        # Convert all cell values to strings
        rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()

    if not rows:
        raise ValueError(f"Excel file is empty: {file_path}")

    # Detect URL column from first row
    first_row = rows[0]
    url_col = _find_url_column(first_row)

    if url_col is not None:
        data_rows = rows[1:]
        start_row = 2
    else:
        url_col = 0
        first_cell = _clean_url(first_row[0]) if first_row else ""
        if LINKEDIN_URL_PATTERN.match(first_cell) or (
            first_cell.startswith("linkedin.com/in/") or first_cell.startswith("www.linkedin.com/in/")
        ):
            data_rows = rows
            start_row = 1
        else:
            data_rows = rows[1:]
            start_row = 2

    for i, row in enumerate(data_rows):
        row_num = start_row + i
        if url_col < len(row):
            url = _validate_url(row[url_col], row_num)
            if url:
                results.append({"url": url, "row": row_num})

    return results


def read_google_sheet(sheet_url: str, credentials_file: Optional[str] = None) -> list[dict]:
    """
    Read LinkedIn URLs from a Google Sheets spreadsheet.

    Args:
        sheet_url: The full Google Sheets URL.
        credentials_file: Path to a Google service account JSON key file.
                         If not provided, reads from GOOGLE_SHEETS_CREDENTIALS env var.
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        raise ImportError(
            "gspread and google-auth are required for Google Sheets. "
            "Install with: pip install gspread google-auth"
        )

    import os

    if credentials_file is None:
        credentials_file = os.getenv("GOOGLE_SHEETS_CREDENTIALS", "credentials.json")

    creds_path = Path(credentials_file)
    if not creds_path.exists():
        raise FileNotFoundError(
            f"Google Sheets credentials file not found: {credentials_file}\n"
            f"Create a service account and download the JSON key from: "
            f"https://console.cloud.google.com/iam-admin/serviceaccounts"
        )

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    credentials = Credentials.from_service_account_file(str(creds_path), scopes=scopes)
    gc = gspread.authorize(credentials)

    # Extract spreadsheet ID from URL
    spreadsheet = gc.open_by_url(sheet_url)
    worksheet = spreadsheet.sheet1  # Use first sheet

    rows = worksheet.get_all_values()
    if not rows:
        raise ValueError(f"Google Sheet is empty: {sheet_url}")

    results = []
    first_row = rows[0]
    url_col = _find_url_column(first_row)

    if url_col is not None:
        data_rows = rows[1:]
        start_row = 2
    else:
        url_col = 0
        test_url = _clean_url(first_row[0]) if first_row else ""
        if LINKEDIN_URL_PATTERN.match(test_url) or test_url.startswith("linkedin.com") or test_url.startswith("www.linkedin.com"):
            data_rows = rows
            start_row = 1
        else:
            data_rows = rows[1:]
            start_row = 2

    for i, row in enumerate(data_rows):
        row_num = start_row + i
        if url_col < len(row):
            url = _validate_url(row[url_col], row_num)
            if url:
                results.append({"url": url, "row": row_num})

    return results


def read_spreadsheet(file_path: str) -> list[dict]:
    """
    Read LinkedIn profile URLs from any supported spreadsheet format.
    Auto-detects the format from the file extension or URL pattern.

    Supported formats:
        - .csv files
        - .xlsx / .xlsm / .xltx / .xltm (Excel) files
        - Google Sheets URLs (https://docs.google.com/spreadsheets/...)

    Args:
        file_path: Path to a local file or a Google Sheets URL.

    Returns:
        List of dicts: [{"url": "https://linkedin.com/in/...", "row": 2}, ...]
    """
    if _is_google_sheets_url(file_path):
        print(f"[INFO] Reading Google Sheet: {file_path}")
        results = read_google_sheet(file_path)
    else:
        path = Path(file_path)
        ext = path.suffix.lower()

        if ext == ".csv":
            print(f"[INFO] Reading CSV file: {file_path}")
            results = read_csv(file_path)
        elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            print(f"[INFO] Reading Excel file: {file_path}")
            results = read_xlsx(file_path)
        else:
            raise ValueError(
                f"Unsupported file format: '{ext}'. "
                f"Supported formats: .csv, .xlsx, or Google Sheets URL."
            )

    print(f"[INFO] Found {len(results)} valid LinkedIn profile URLs.")
    return results
